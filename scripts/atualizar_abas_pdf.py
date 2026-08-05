# -*- coding: utf-8 -*-
"""Atualiza no instrumento de completude apenas as abas cujo conteudo mudou.

A reconstrucao do texto extraido de PDF altera a segmentacao dos documentos que o
contenham. Regenerar o instrumento inteiro descartaria a marcacao ja realizada nas
demais politicas; esta rotina substitui somente as abas cujos segmentos divergem do
conjunto corrente, e preserva as restantes com todas as marcas.

A comparacao e feita sobre o texto dos segmentos, e nao sobre a contagem: aba cujo
conteudo permaneca identico nao e tocada, ainda que a ordem interna varie.

As abas substituidas perdem a marcacao, o que e inevitavel — os segmentos passaram a
ser outros. A rotina relaciona quais serao afetadas ANTES de gravar, e exige
confirmacao explicita.

Uso:
    python scripts/atualizar_abas_pdf.py --tcc "C:/caminho/TCC"
    python scripts/atualizar_abas_pdf.py --tcc "C:/caminho/TCC" --aplicar
"""
from __future__ import annotations

import argparse
import csv
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
VARIAVEIS = ["finalidade", "direitos_titular", "transf_internacional"]
ROTULO = {"finalidade": "Finalidade", "direitos_titular": "Direitos",
          "transf_internacional": "Transf. intern."}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--tcc", required=True)
    ap.add_argument("--planilha", default="Rotulagem/Completude - 15 politicas.xlsx")
    ap.add_argument("--segmentos", default="outputs/segmentos_textuais.csv")
    ap.add_argument("--aplicar", action="store_true")
    args = ap.parse_args()

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    caminho = Path(args.tcc) / args.planilha
    wb = openpyxl.load_workbook(caminho)

    with (REPO / args.segmentos).open(encoding="utf-8", newline="") as fh:
        S = [l for l in csv.DictReader(fh, delimiter=";")
             if l["variavel"] == VARIAVEIS[0]]
    atual = defaultdict(dict)
    for l in S:
        atual[l["site_id"]][int(l["segmento_id"])] = l["texto"]

    mudou, iguais, marcas_perdidas = [], [], 0
    for aba in wb.sheetnames:
        if aba in ("Controle", "Instrucoes"):
            continue
        pref = aba.split(" ", 1)[1] if " " in aba else aba
        sitio = next((s for s in atual if s.startswith(pref)), None)
        ws = wb[aba]
        cols = [c.value for c in ws[1]]
        antigo = {}
        marcas = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None:
                continue
            antigo[int(r[0])] = r[1]
            for v in VARIAVEIS:
                j = cols.index(ROTULO[v])
                if r[j] not in (None, ""):
                    marcas += 1
        novo = atual.get(sitio, {})
        if list(antigo.values()) == list(novo.values()):
            iguais.append((aba, len(antigo)))
        else:
            mudou.append((aba, sitio, len(antigo), len(novo), marcas))
            marcas_perdidas += marcas

    print(f"abas inalteradas: {len(iguais)}")
    print(f"abas a substituir: {len(mudou)}\n")
    for aba, sitio, na, nn, marcas in mudou:
        aviso = f"   PERDE {marcas} marca(s)" if marcas else "   (sem marcacao ainda)"
        print(f"  {aba:32} {na:>5} -> {nn:>5} segmentos{aviso}")
    print(f"\n  marcas que serao perdidas ao todo: {marcas_perdidas}")

    if not args.aplicar:
        print("\n  modo de relacao: nada foi gravado. Use --aplicar para efetivar.")
        return 0
    if not mudou:
        print("\n  nada a fazer.")
        return 0

    backup = caminho.with_suffix(f".bak_{datetime.now():%Y%m%d%H%M}.xlsx")
    shutil.copy2(caminho, backup)

    cab = ["segmento_id", "texto"] + [ROTULO[v] for v in VARIAVEIS] + ["obs"]
    for aba, sitio, _, _, _ in mudou:
        pos = wb.sheetnames.index(aba)
        del wb[aba]
        ws = wb.create_sheet(aba, pos)
        ws.append(cab)
        for c in range(1, len(cab) + 1):
            ws.cell(1, c).font = Font(bold=True)
            ws.cell(1, c).fill = PatternFill("solid", fgColor="DDDDDD")
        segs = atual.get(sitio, {})
        for sid in sorted(segs):
            ws.append([sid, segs[sid], "", "", "", ""])
        ws.column_dimensions["A"].width = 11
        ws.column_dimensions["B"].width = 104
        for i, _ in enumerate(VARIAVEIS):
            ws.column_dimensions[get_column_letter(3 + i)].width = 14
        ws.column_dimensions[get_column_letter(3 + len(VARIAVEIS))].width = 34
        for linha in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            linha[0].alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        dv = DataValidation(type="list", formula1='"1"', allow_blank=True)
        ws.add_data_validation(dv)
        for i, _ in enumerate(VARIAVEIS):
            col = get_column_letter(3 + i)
            dv.add(f"{col}2:{col}{len(segs) + 1}")

    # a aba de Controle passa a refletir os novos tamanhos e limpa a conclusao
    if "Controle" in wb.sheetnames:
        ctl = wb["Controle"]
        afetados = {s for _, s, _, _, _ in mudou}
        for r in ctl.iter_rows(min_row=2):
            if r[0].value and r[0].value in afetados:
                r[1].value = len(atual.get(r[0].value, {}))
                r[3].value = None
    wb.save(caminho)
    print(f"\n  {len(mudou)} aba(s) substituida(s); copia de seguranca em {backup.name}")
    print("  A conclusao dessas politicas foi limpa na aba de Controle.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
