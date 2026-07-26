# -*- coding: utf-8 -*-
"""Concordancia entre avaliadores no subconjunto cego.

MEDIDAS REPORTADAS
------------------
O coeficiente kappa de Cohen (1960) corrige a concordancia observada pela
concordancia esperada ao acaso. A correcao, contudo, torna o coeficiente sensivel
a distribuicao marginal: Feinstein e Cicchetti (1990) descreveram o paradoxo pelo
qual concordancia observada elevada convive com kappa baixo quando a prevalencia e
desequilibrada, situacao esperada em variaveis como transferencia internacional.

Reportam-se por isso quatro grandezas conjuntamente:

  - concordancia observada, a proporcao bruta de coincidencias;
  - kappa de Cohen, com intervalo por reamostragem;
  - indice de prevalencia, a diferenca absoluta entre as proporcoes das duas
    respostas concordantes, que dispara o primeiro paradoxo quando se aproxima
    da unidade;
  - indice de vies, a diferenca absoluta entre as marginais dos avaliadores, que
    dispara o segundo paradoxo;
  - kappa ajustado por prevalencia e vies, que remove ambos os efeitos e serve de
    contraponto ao kappa bruto.

A interpretacao qualitativa segue as faixas de Landis e Koch (1977). Reporta-se a
faixa junto do valor, jamais em substituicao a ele.

TRATAMENTO DA AVALIABILIDADE
----------------------------
A discordancia sobre SE o sitio e avaliavel difere da discordancia sobre O QUE ele
apresenta. A primeira e computada em separado, sobre a variavel de status.
Misturar as duas confundiria divergencia de julgamento com divergencia sobre a
qualidade da captura.

POPULACAO DE CADA VARIAVEL
--------------------------
A concordancia e apurada sobre a mesma populacao em que a variavel e modelada, sob
pena de medir coisa diversa daquela que sustenta o classificador.

O canal do titular nao constitui propriedade do texto da politica: pode ser
divulgado em pagina de contato ou no rodape. Apura-se, por isso, sobre todos os
sitios em que ambos os avaliadores atribuiram valor.

As tres variaveis textuais so existem enquanto propriedade de um texto. Nos sitios
sem politica elas valem zero por definicao, e a coincidencia entre avaliadores e
automatica. Incluir esses sitios elevaria a concordancia sem que julgamento algum
tivesse sido exercido. Restringe-se, portanto, aos sitios que AMBOS julgaram
dotados de texto avaliavel.

Uso:
    python scripts/calcular_kappa.py --segundo "caminho/planilha_preenchida.xlsx"
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path

import numpy as np

REPO = Path(__file__).resolve().parents[1]
VARIAVEIS = ["tem_canal_titular", "finalidade", "direitos_titular", "transf_internacional"]

FAIXAS = [(0.81, "quase perfeita"), (0.61, "substancial"), (0.41, "moderada"),
          (0.21, "razoavel"), (0.0, "leve"), (-1.0, "pobre")]


def faixa(k):
    for corte, nome in FAIXAS:
        if k >= corte:
            return nome
    return "pobre"


def kappa_cohen(a, b):
    a = np.asarray(a); b = np.asarray(b)
    n = len(a)
    if n == 0:
        return float("nan"), float("nan"), float("nan"), float("nan")
    po = float(np.mean(a == b))
    cats = sorted(set(a) | set(b))
    pe = sum((np.mean(a == c)) * (np.mean(b == c)) for c in cats)
    k = (po - pe) / (1 - pe) if pe < 1 else 1.0
    # indices de prevalencia e vies, definidos para o caso binario
    if set(cats) <= {0, 1}:
        n11 = int(np.sum((a == 1) & (b == 1))); n00 = int(np.sum((a == 0) & (b == 0)))
        n10 = int(np.sum((a == 1) & (b == 0))); n01 = int(np.sum((a == 0) & (b == 1)))
        ip = abs(n11 - n00) / n
        iv = abs(n10 - n01) / n
        pabak = 2 * po - 1
    else:
        ip = iv = pabak = float("nan")
    return k, po, ip, iv, pabak


def ic_reamostragem(a, b, reps=5000, semente=20260720):
    rng = np.random.default_rng(semente)
    a = np.asarray(a); b = np.asarray(b); n = len(a)
    ks = []
    for _ in range(reps):
        i = rng.integers(0, n, n)
        r = kappa_cohen(a[i], b[i])[0]
        if np.isfinite(r):
            ks.append(r)
    if not ks:
        return float("nan"), float("nan")
    return float(np.percentile(ks, 2.5)), float(np.percentile(ks, 97.5))


def le_devolutiva(caminho, aba=None):
    """Le a devolutiva do segundo avaliador em planilha ou em texto separado.

    O avaliador pode devolver o material em formato diverso do entregue; aceitam-se
    ambos, com deteccao do separador para evitar que a escolha do editor altere a
    leitura.
    """
    caminho = str(caminho)
    if caminho.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(caminho, data_only=True)
        ws = wb[aba] if aba else wb[wb.sheetnames[0]]
        cols = [c.value for c in ws[1]]
        return [{cols[j]: ws.cell(r, j + 1).value for j in range(len(cols))}
                for r in range(2, ws.max_row + 1)]
    with open(caminho, encoding="utf-8-sig", newline="") as fh:
        amostra = fh.read(8192)
        fh.seek(0)
        sep = ";" if amostra.count(";") >= amostra.count(",") else ","
        return list(csv.DictReader(fh, delimiter=sep))


def norm(v):
    if v is None or str(v).strip() == "":
        return None
    s = str(v).strip().upper()
    if s in ("NA", "N/A"):
        return "NA"
    if s in ("0", "0.0"):
        return 0
    if s in ("1", "1.0"):
        return 1
    return s.lower()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--primario", default="rotulagem_b9.csv")
    ap.add_argument("--segundo", required=True, help="planilha preenchida pelo segundo avaliador")
    ap.add_argument("--out", default="outputs/kappa_resultados.csv")
    args = ap.parse_args()

    with (REPO / args.primario).open(encoding="utf-8-sig", newline="") as fh:
        P = {r["site_id"]: r for r in csv.DictReader(fh, delimiter=";")}
    S = {r["site_id"]: r for r in le_devolutiva(args.segundo) if r.get("site_id")}

    comuns = [s for s in S if s in P]
    print(f"sitios na planilha do segundo avaliador: {len(S)}")
    print(f"sitios pareados com a rotulagem primaria: {len(comuns)}")
    faltantes = [s for s in S if s not in P]
    if faltantes:
        print(f"  ATENCAO: {len(faltantes)} sem par: {faltantes[:5]}")
    vazios = [s for s in comuns if norm(S[s].get("status")) is None]
    if vazios:
        print(f"  ATENCAO: {len(vazios)} sitios sem status preenchido; serao ignorados")
    comuns = [s for s in comuns if norm(S[s].get("status")) is not None]
    print(f"sitios efetivamente comparados: {len(comuns)}\n")

    linhas = []

    # --- avaliabilidade ---
    a = [1 if norm(P[s].get("status")) == "text" else 0 for s in comuns]
    b = [1 if norm(S[s].get("status")) == "text" else 0 for s in comuns]
    k, po, ip, iv, pabak = kappa_cohen(a, b)
    lo, hi = ic_reamostragem(a, b)
    print("=" * 100)
    print("AVALIABILIDADE  (o sitio apresenta politica com corpo avaliavel?)")
    print("=" * 100)
    print(f"  n={len(comuns)}  concordancia observada={po*100:.1f}%  kappa={k:.3f} "
          f"[{lo:.3f}, {hi:.3f}]  ({faixa(k)})")
    print(f"  indice de prevalencia={ip:.3f}   indice de vies={iv:.3f}   PABAK={pabak:.3f}")
    linhas.append({"variavel": "status_avaliavel", "n": len(comuns), "concordancia": po,
                   "kappa": k, "ic_inf": lo, "ic_sup": hi, "ind_prevalencia": ip,
                   "ind_vies": iv, "pabak": pabak, "faixa": faixa(k)})

    # --- variaveis substantivas ---
    com_texto = [s for s in comuns
                 if norm(P[s].get("status")) == "text" and norm(S[s].get("status")) == "text"]
    print("\n" + "=" * 100)
    print("VARIAVEIS")
    print("=" * 100)
    print(f"  canal do titular: todos os {len(comuns)} sitios com valor atribuido")
    print(f"  variaveis textuais: os {len(com_texto)} sitios que ambos julgaram dotados de texto")
    print()
    print(f"  {'variavel':24}{'n':>5}{'concord.':>10}{'kappa':>9}{'IC 95%':>18}"
          f"{'i.prev':>8}{'i.vies':>8}{'PABAK':>8}  faixa")
    for v in VARIAVEIS:
        base = comuns if v == "tem_canal_titular" else com_texto
        pares = [(norm(P[s].get(v)), norm(S[s].get(v))) for s in base]
        val = [(x, y) for x, y in pares if x in (0, 1) and y in (0, 1)]
        if len(val) < 5:
            print(f"  {v:24}{len(val):>5}   dados insuficientes")
            continue
        a = [x for x, _ in val]; b = [y for _, y in val]
        k, po, ip, iv, pabak = kappa_cohen(a, b)
        lo, hi = ic_reamostragem(a, b)
        print(f"  {v:24}{len(val):>5}{po*100:>9.1f}%{k:>9.3f}"
              f"{lo:>9.3f},{hi:>7.3f}{ip:>8.3f}{iv:>8.3f}{pabak:>8.3f}  {faixa(k)}")
        linhas.append({"variavel": v, "n": len(val), "concordancia": po, "kappa": k,
                       "ic_inf": lo, "ic_sup": hi, "ind_prevalencia": ip,
                       "ind_vies": iv, "pabak": pabak, "faixa": faixa(k)})

    print("\n" + "=" * 100)
    print("LEITURA")
    print("=" * 100)
    for r in linhas:
        lacuna = r["pabak"] - r["kappa"]
        print(f"  {r['variavel']}: concordancia {r['concordancia']*100:.1f}%, "
              f"kappa {r['kappa']:.3f} ({r['faixa']})")
        if lacuna > 0.06:
            print(f"    O kappa fica {lacuna:.3f} abaixo do ajustado por prevalencia e vies, "
                  f"com indice de prevalencia de {r['ind_prevalencia']:.2f}.")
            print("    A defasagem decorre do desequilibrio entre as respostas, e nao de")
            print("    divergencia de julgamento — situacao descrita por Feinstein e Cicchetti")
            print("    (1990). Reportar as duas medidas em conjunto, jamais o kappa isolado.")
        if r["ind_vies"] > 0.15:
            print(f"    Indice de vies de {r['ind_vies']:.2f}: os avaliadores aplicam o criterio")
            print("    com severidade distinta. Convem examinar a direcao das discordancias.")
        largura = r["ic_sup"] - r["ic_inf"]
        if largura > 0.35:
            print(f"    Intervalo de {largura:.2f} de largura: com n = {r['n']} a estimativa e")
            print("    imprecisa, e o valor pontual nao deve ser lido sem a faixa.")
    pior = min((r for r in linhas if r["variavel"] != "status_avaliavel"),
               key=lambda x: x["kappa"], default=None)
    if pior:
        print(f"\n  A concordancia entre avaliadores constitui teto pratico para o")
        print(f"  classificador: nao se espera que um modelo distinga o que dois leitores")
        print(f"  humanos, aplicando o mesmo codebook, classificam de modo divergente.")
        print(f"  O teto mais baixo cabe a variavel {pior['variavel']}, com kappa "
              f"{pior['kappa']:.3f}.")

    out = REPO / args.out
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(linhas[0].keys()), delimiter=";")
        w.writeheader(); w.writerows(linhas)
    print(f"\nsaida: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
