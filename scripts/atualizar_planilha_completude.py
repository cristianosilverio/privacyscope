# -*- coding: utf-8 -*-
"""Transporta a marcacao exaustiva para o conjunto de segmentos reconstruido.

MOTIVO
------
A etapa de preparo do texto mudou em dois pontos: o material das politicas
publicadas em PDF voltou a ser incorporado, e a remocao por repeticao deu lugar a
deduplicacao por sitio. As duas alteracoes mexem no conjunto de segmentos, e a
marcacao exaustiva das quinze politicas foi feita sobre o conjunto anterior.

Refazer a marcacao inteira seria desperdicio e, pior, risco: a marcacao original foi
conduzida numa sessao, sob leitura continua do codebook, e reproduzi-la em outra
sessao introduziria variacao de criterio onde hoje nao ha nenhuma.

O TRANSPORTE E POR TEXTO, E NAO POR POSICAO
-------------------------------------------
O identificador de segmento e o indice da sentenca dentro do documento reconstituido,
e portanto se desloca quando material novo entra no meio. Transportar por
identificador deslocaria as marcas junto, atribuindo silenciosamente a marca de uma
sentenca a outra.

O transporte se faz por igualdade EXATA do texto, dentro do mesmo sitio. Sentenca cujo
texto ja foi julgado conserva o julgamento; sentenca cujo texto nao consta da marcacao
anterior e nova e vai assinalada para julgamento.

Quando a marcacao anterior traz o mesmo texto em mais de uma linha — situacao corrente,
porque a deduplicacao ainda nao existia — e as linhas divergem, prevalece a marca
POSITIVA. A divergencia decorria de o intervalo da transcricao cobrir uma ocorrencia e
nao as outras, e nao de julgamento distinto do anotador.

IDENTIFICADOR DO SITIO
----------------------
O nome da aba NAO serve de identificador: o Excel limita nomes de planilha a 31
caracteres, e ao menos uma politica da amostra tem dominio que excede esse limite,
tendo sido truncada para `primeirainfanciaemdados.or`. O identificador provem da aba
`Controle`, que registra o dominio por extenso, e a correspondencia com as abas se faz
pela ordem em que ambas foram geradas.

SAIDA
-----
Grava-se ARQUIVO NOVO. A planilha original permanece intacta, e continua sendo o
registro do que foi julgado sobre o conjunto anterior.

Uso:
    python scripts/atualizar_planilha_completude.py
    python scripts/atualizar_planilha_completude.py --tcc "C:/caminho/TCC"
"""
from __future__ import annotations

import argparse
import csv
import os
from collections import defaultdict
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
VARIAVEIS = ["finalidade", "direitos_titular", "transf_internacional"]
COLUNAS = ["Finalidade", "Direitos", "Transf. intern."]
CABECALHO = ["segmento_id", "texto"] + COLUNAS + ["obs"]
MARCA_NOVO = "NOVO - MARCAR"


def le_planilha(caminho):
    """Marcas da planilha vigente, consolidadas por texto dentro de cada sitio."""
    import openpyxl
    wb = openpyxl.load_workbook(caminho, data_only=True)
    if "Controle" not in wb.sheetnames:
        raise SystemExit("ABORTADO: a planilha nao tem aba `Controle`, de onde provem "
                         "o dominio por extenso de cada politica.")
    dominios = [str(r[0]).strip() for r in wb["Controle"].iter_rows(min_row=2, values_only=True)
                if r and r[0] and str(r[0]).strip() not in ("TOTAL", "None")]
    abas = [n for n in wb.sheetnames if n not in ("Controle", "Instrucoes")]
    if len(abas) != len(dominios):
        raise SystemExit(f"ABORTADO: {len(abas)} abas de politica contra "
                         f"{len(dominios)} linhas em `Controle`; a correspondencia "
                         f"por ordem nao e segura.")
    for aba, dom in zip(abas, dominios):
        # A aba pode estar truncada, mas o prefixo tem de bater: se nao bater, a
        # ordem das duas listas divergiu e o transporte atribuiria marcas ao sitio
        # errado, sem que nada o denuncie.
        corpo = aba.split(" ", 1)[1] if " " in aba else aba
        if not dom.startswith(corpo.rstrip(".")):
            raise SystemExit(f"ABORTADO: a aba `{aba}` nao corresponde a `{dom}`.")

    marcas, obs, ordem = defaultdict(dict), defaultdict(dict), {}
    for aba, dom in zip(abas, dominios):
        ordem[dom] = aba
        for r in wb[aba].iter_rows(min_row=2, values_only=True):
            if not r or r[1] is None:
                continue
            texto = str(r[1])
            d = marcas[dom].setdefault(texto, {c: None for c in COLUNAS})
            for k, c in enumerate(COLUNAS):
                if len(r) > 2 + k and str(r[2 + k]).strip() == "1":
                    d[c] = "1"                      # positiva prevalece
            if len(r) > 5 and r[5]:
                obs[dom][texto] = str(r[5])
    inst = None
    if "Instrucoes" in wb.sheetnames:
        inst = [[c for c in row] for row in wb["Instrucoes"].iter_rows(values_only=True)]
    return dominios, ordem, marcas, obs, inst


def le_segmentos(caminho, sitios):
    """Segmentos do conjunto reconstruido, em ordem de documento."""
    csv.field_size_limit(10 ** 7)
    por_sitio = defaultdict(list)
    with caminho.open(encoding="utf-8", newline="") as fh:
        leitor = csv.DictReader(fh, delimiter=";")
        for r in leitor:
            if r["variavel"] != VARIAVEIS[0] or r["site_id"] not in sitios:
                continue
            if r.get("duplicata") == "1":
                continue
            por_sitio[r["site_id"]].append((int(r["segmento_id"]), r["texto"]))
    for s in por_sitio:
        por_sitio[s].sort()
    return por_sitio


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--tcc", default=os.environ.get("PRIVACYSCOPE_TCC"))
    ap.add_argument("--planilha", default="Rotulagem/Completude - 15 politicas.xlsx")
    ap.add_argument("--segmentos", default="outputs/segmentos_textuais.csv")
    ap.add_argument("--saida",
                    default="Rotulagem/Completude - 15 politicas (v2).xlsx")
    args = ap.parse_args()
    if not args.tcc:
        print("ERRO: informe --tcc ou defina PRIVACYSCOPE_TCC.")
        return 2

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    tcc = Path(args.tcc)
    dominios, ordem, marcas, obs, inst = le_planilha(tcc / args.planilha)
    novos = le_segmentos(REPO / args.segmentos, set(dominios))
    ausentes = [d for d in dominios if d not in novos]
    if ausentes:
        raise SystemExit(f"ABORTADO: sem segmentos para {ausentes}. O conjunto "
                         f"reconstruido nao cobre todas as politicas marcadas.")

    wb = openpyxl.Workbook()
    ctrl = wb.active
    ctrl.title = "Controle"
    ctrl.append(["Política", "Segmentos", "Já julgados", "Novos a marcar", "Concluída (sim)"])
    if inst:
        ws = wb.create_sheet("Instrucoes")
        for row in inst:
            ws.append(list(row))

    negrito = Font(bold=True)
    realce = PatternFill("solid", fgColor="FFF2CC")
    total_novos = total_seg = 0
    perdidos = {}

    for dom in dominios:
        ws = wb.create_sheet(ordem[dom][:31])
        ws.append(CABECALHO)
        for c in ws[1]:
            c.font = negrito
        vistos, n_novos = set(), 0
        for sid, texto in novos[dom]:
            vistos.add(texto)
            d = marcas[dom].get(texto)
            if d is None:
                n_novos += 1
                ws.append([sid, texto, None, None, None, MARCA_NOVO])
                for c in ws[ws.max_row]:
                    c.fill = realce
            else:
                ws.append([sid, texto, d["Finalidade"], d["Direitos"],
                           d["Transf. intern."], obs[dom].get(texto)])
        # Texto julgado que nao reaparece: nao e erro por si, mas tem de ser visivel.
        sumiu = [t for t in marcas[dom] if t not in vistos]
        com_marca = [t for t in sumiu if any(marcas[dom][t].values())]
        if sumiu:
            perdidos[dom] = (len(sumiu), len(com_marca))

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:F{ws.max_row}"
        for col, larg in zip("ABCDEF", (12, 110, 12, 12, 16, 18)):
            ws.column_dimensions[col].width = larg
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            row[0].alignment = Alignment(wrap_text=False, vertical="top")
        total_seg += len(novos[dom]); total_novos += n_novos
        ctrl.append([dom, len(novos[dom]), len(novos[dom]) - n_novos, n_novos,
                     "" if n_novos else "sim"])

    ctrl.append([]); ctrl.append(["TOTAL", total_seg, total_seg - total_novos, total_novos, ""])
    for c in ctrl[1]:
        c.font = negrito
    ctrl.freeze_panes = "A2"
    for col, larg in zip("ABCDE", (34, 12, 13, 16, 16)):
        ctrl.column_dimensions[col].width = larg

    destino = tcc / args.saida
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)

    print(f"planilha vigente: {args.planilha}")
    print(f"conjunto novo:    {args.segmentos}\n")
    print(f"  {'politica':32}{'segmentos':>11}{'julgados':>10}{'A MARCAR':>10}")
    for row in ctrl.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0] != "TOTAL":
            print(f"  {row[0]:32}{row[1]:>11}{row[2]:>10}{row[3]:>10}")
    print(f"  {'TOTAL':32}{total_seg:>11}{total_seg - total_novos:>10}{total_novos:>10}")
    if perdidos:
        print(f"\n  textos julgados que nao reaparecem no conjunto novo:")
        for dom, (n, m) in sorted(perdidos.items()):
            print(f"    {dom:32}{n:>5} texto(s), {m} com marca positiva")
    else:
        print(f"\n  nenhum texto julgado desapareceu do conjunto novo.")
    print(f"\nsaida: {destino}")
    print(f"As linhas a marcar estao realcadas e trazem `{MARCA_NOVO}` na coluna obs;")
    print(f"o filtro do cabecalho isola-as em cada aba.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
