# -*- coding: utf-8 -*-
"""Instrumento de validacao da completude da anotacao de trechos.

O QUE SE MEDE
-------------
A derivacao de rotulos de segmento a partir das transcricoes apoia-se em uma
suposicao verificavel: a de que a passagem transcrita esgota, ou quase esgota, o
conteudo relevante do documento. Se documentos positivos contiverem muitas outras
passagens igualmente relevantes que nao foram transcritas, tratar os demais
segmentos como nao rotulados e imprescindivel — e a fracao nao rotulada sera
elevada. Se as passagens adicionais forem raras, a anotacao e praticamente
completa e o conjunto derivado e solido.

A grandeza de interesse e, portanto, o numero de segmentos relevantes por documento
que a transcricao original NAO alcancou.

DESENHO
-------
A rotulagem original respondeu "esta variavel esta presente?" e transcreveu UMA
passagem justificadora. Este instrumento formula pergunta distinta: "quais de todos
estes segmentos sao relevantes para cada variavel?". O anotador percorre a
totalidade dos segmentos de um numero reduzido de politicas e assinala todos os
relevantes.

A transcricao original NAO e exibida. Exibi-la ancoraria o julgamento na passagem
ja escolhida e produziria concordancia artificial, esvaziando a medida.

O sorteio e estratificado pelas tres variaveis, de modo que cada uma comparece em
documentos positivos, e utiliza semente declarada.

Esta tarefa cabe ao anotador primario, e nao ao segundo avaliador: mede-se a
completude da anotacao propria, nao a concordancia entre avaliadores, que e objeto
do subconjunto cego.

Uso:
    python scripts/preparar_validacao_completude.py --tcc "C:/caminho/TCC"
"""
from __future__ import annotations

import argparse
import csv
import random
from collections import defaultdict
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
VARIAVEIS = ["finalidade", "direitos_titular", "transf_internacional"]
ROTULO = {"finalidade": "Finalidade", "direitos_titular": "Direitos",
          "transf_internacional": "Transf. intern."}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--segmentos", default="outputs/segmentos_textuais.csv")
    ap.add_argument("--rotulagem", default="rotulagem_b9.csv")
    ap.add_argument("--tcc", required=True)
    ap.add_argument("--politicas", type=int, default=15)
    ap.add_argument("--semente", type=int, default=20260725)
    args = ap.parse_args()

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    with (REPO / args.segmentos).open(encoding="utf-8", newline="") as fh:
        S = list(csv.DictReader(fh, delimiter=";"))
    with (REPO / args.rotulagem).open(encoding="utf-8-sig", newline="") as fh:
        R = {r["site_id"]: r for r in csv.DictReader(fh, delimiter=";")}

    # candidatos: documentos positivos em ao menos uma variavel
    positivos = defaultdict(set)
    for v in VARIAVEIS:
        for s, r in R.items():
            if r.get("status") == "text" and r.get(v) == "1":
                positivos[v].add(s)

    rng = random.Random(args.semente)
    escolhidos = []
    # cota por variavel, priorizando a de menor prevalencia
    for v in sorted(VARIAVEIS, key=lambda x: len(positivos[x])):
        cota = max(1, args.politicas // len(VARIAVEIS))
        disp = sorted(positivos[v] - set(escolhidos))
        rng.shuffle(disp)
        escolhidos.extend(disp[:cota])
    disp = sorted(set().union(*positivos.values()) - set(escolhidos))
    rng.shuffle(disp)
    escolhidos.extend(disp[: max(0, args.politicas - len(escolhidos))])
    escolhidos = escolhidos[: args.politicas]

    print(f"politicas sorteadas: {len(escolhidos)} (semente {args.semente})")
    for v in VARIAVEIS:
        print(f"  positivas em {ROTULO[v]:16}: "
              f"{sum(1 for s in escolhidos if R[s].get(v) == '1')}")

    # segmentos, sem duplicar entre variaveis
    por_sitio = defaultdict(dict)
    for l in S:
        if l["site_id"] in escolhidos:
            por_sitio[l["site_id"]][int(l["segmento_id"])] = l["texto"]

    # Uma aba por politica. Percorrer sete mil linhas numa aba unica dificulta
    # retomar o trabalho e obscurece o progresso; a separacao permite concluir uma
    # politica por sessao, o que preserva a atencao ao longo da tarefa.
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    controle = wb.active; controle.title = "Controle"
    controle.append(["Política", "Segmentos", "Positiva em", "Concluída (sim)"])
    for c in range(1, 5):
        controle.cell(1, c).font = Font(bold=True)
        controle.cell(1, c).fill = PatternFill("solid", fgColor="DDDDDD")

    cab = ["segmento_id", "texto"] + [ROTULO[v] for v in VARIAVEIS] + ["obs"]
    total = 0
    for ordem, sitio in enumerate(sorted(escolhidos, key=lambda s: len(por_sitio[s])), 1):
        aba = f"{ordem:02d} {sitio[:26]}"
        ws = wb.create_sheet(aba)
        ws.append(cab)
        for c in range(1, len(cab) + 1):
            ws.cell(1, c).font = Font(bold=True)
            ws.cell(1, c).fill = PatternFill("solid", fgColor="DDDDDD")
        n = 0
        for sid in sorted(por_sitio[sitio]):
            ws.append([sid, por_sitio[sitio][sid], "", "", "", ""])
            n += 1; total += 1
        ws.column_dimensions["A"].width = 11
        ws.column_dimensions["B"].width = 104
        for i, _ in enumerate(VARIAVEIS):
            ws.column_dimensions[get_column_letter(3 + i)].width = 14
        ws.column_dimensions[get_column_letter(3 + len(VARIAVEIS))].width = 34
        for linha in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            linha[0].alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        # lista suspensa nas colunas de marcacao, para dispensar digitacao
        dv = DataValidation(type="list", formula1='"1"', allow_blank=True)
        ws.add_data_validation(dv)
        for i, _ in enumerate(VARIAVEIS):
            col = get_column_letter(3 + i)
            dv.add(f"{col}2:{col}{n + 1}")
        pos = ", ".join(ROTULO[v] for v in VARIAVEIS if R[sitio].get(v) == "1")
        controle.append([sitio, n, pos or "—", ""])

    controle.append([])
    controle.append(["TOTAL", total, "", ""])
    controle.cell(controle.max_row, 1).font = Font(bold=True)
    controle.column_dimensions["A"].width = 30
    controle.column_dimensions["B"].width = 12
    controle.column_dimensions["C"].width = 42
    controle.column_dimensions["D"].width = 18
    controle.freeze_panes = "A2"

    wi = wb.create_sheet("Instrucoes", 1)
    for t in [
        "Validacao da completude da anotacao de trechos",
        "",
        f"Politicas: {len(escolhidos)}   Segmentos: {total}   Semente: {args.semente}",
        "",
        "OBJETIVO",
        "Medir quantas passagens relevantes existem alem daquela transcrita na rotulagem",
        "original. A rotulagem respondeu 'a variavel esta presente?' e registrou UMA",
        "passagem. Aqui a pergunta e outra: quais de TODOS estes segmentos sao relevantes.",
        "",
        "COMO PREENCHER",
        "Ha uma aba por politica, ordenadas da menor para a maior, e uma aba de Controle",
        "com o tamanho de cada uma. Percorra os segmentos e marque 1 nas colunas das",
        "variaveis para as quais o segmento e relevante. Deixe em branco quando nao for.",
        "Um segmento pode ser relevante para mais de uma variavel, e a maioria nao sera",
        "relevante para nenhuma.",
        "",
        "As tres primeiras abas somam menos de cem segmentos e servem para calibrar o",
        "ritmo. A ultima concentra um terco do total: convem reserva-la para sessao",
        "propria. Registre na aba de Controle a conclusao de cada politica, para poder",
        "interromper e retomar sem perder o ponto.",
        "",
        "REGRAS",
        "1. NAO consulte a rotulagem original nem as colunas de evidencia. A transcricao",
        "   anterior nao e exibida de proposito: ve-la ancoraria o julgamento e produziria",
        "   concordancia artificial.",
        "2. Aplique o mesmo codebook da rotulagem original.",
        "3. Marque TODAS as passagens relevantes, e nao apenas a mais representativa.",
        "4. Nao consulte os sitios ao vivo: os segmentos vem do material coletado.",
        "5. Use obs para registrar duvida ou caso limitrofe.",
        "",
        "OBSERVACAO",
        "Muitos segmentos serao material de navegacao ou rodape. Descarte-os rapidamente;",
        "a leitura atenta cabe aos segmentos de conteudo.",
    ]:
        wi.append([t])
    wi.column_dimensions["A"].width = 92

    destino = Path(args.tcc) / "Rotulagem" / f"Completude - {len(escolhidos)} politicas.xlsx"
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    print(f"\nsegmentos a percorrer: {total} "
          f"(mediana de {total // max(len(escolhidos),1)} por politica)")
    print(f"instrumento: {destino}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
