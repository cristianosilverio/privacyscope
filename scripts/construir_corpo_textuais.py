# -*- coding: utf-8 -*-
"""Extrai o corpo rotulado das variaveis textuais do instrumento de completude.

O conjunto de treino das tres variaveis textuais e a marcacao exaustiva das quinze
politicas: em cada uma delas o anotador percorreu a totalidade dos segmentos e
assinalou os relevantes, de sorte que a ausencia de marca constitui negativo
verificado, e nao presumido.

A unidade e a SENTENCA. A supervisao a distancia a partir das passagens transcritas
durante a rotulagem foi avaliada e descartada: sua revocacao, aferida contra esta
mesma marcacao, situa-se entre 8% e 24%, porque a rotulagem original registrava uma
passagem justificadora e nao a totalidade delas.

Uso:
    python scripts/construir_corpo_textuais.py --tcc "C:/caminho/TCC"
"""
from __future__ import annotations

import argparse
import csv
import os
from collections import Counter
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
VARIAVEIS = ["finalidade", "direitos_titular", "transf_internacional"]
ROTULO = {"finalidade": "Finalidade", "direitos_titular": "Direitos",
          "transf_internacional": "Transf. intern."}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--tcc", default=os.environ.get("PRIVACYSCOPE_TCC"),
                    help="pasta raiz do TCC; na ausencia do argumento adota-se a variavel de ambiente PRIVACYSCOPE_TCC")
    ap.add_argument("--planilha", default="Rotulagem/Completude - 15 politicas.xlsx")
    ap.add_argument("--out", default="outputs/segmentos_rotulados.csv")
    args = ap.parse_args()

    if not args.tcc:
        print("ERRO: a pasta do TCC nao foi informada.")
        print("  Informe --tcc, ou defina a variavel de ambiente:")
        print("    PowerShell:  $env:PRIVACYSCOPE_TCC = \"C:\\caminho\\TCC\"")
        print("    bash:        export PRIVACYSCOPE_TCC=/caminho/TCC")
        return 2

    import openpyxl

    wb = openpyxl.load_workbook(Path(args.tcc) / args.planilha, data_only=True)

    # O identificador do sitio NAO provem do nome da aba. O Excel limita nomes de
    # planilha a 31 caracteres, e ao menos um dominio da amostra excede esse limite,
    # tendo sido gravado como `primeirainfanciaemdados.or`; derivar dali produz
    # identificador truncado, que rompe em silencio qualquer juncao posterior. A aba
    # `Controle` registra o dominio por extenso, e a correspondencia se faz pela ordem
    # em que as duas listas foram geradas.
    if "Controle" not in wb.sheetnames:
        raise SystemExit("ABORTADO: a planilha nao tem aba `Controle`.")
    ctrl = list(wb["Controle"].iter_rows(min_row=1, values_only=True))
    cab = [str(c).strip() if c is not None else "" for c in ctrl[0]]
    j_conc = next((k for k, c in enumerate(cab) if c.lower().startswith("conclu")), None)
    dominios, concluidas = [], set()
    for r in ctrl[1:]:
        if not r or not r[0] or str(r[0]).strip() in ("TOTAL", "None"):
            continue
        dom = str(r[0]).strip()
        dominios.append(dom)
        if (j_conc is not None and len(r) > j_conc and r[j_conc]
                and str(r[j_conc]).strip().lower() in ("sim", "s", "x", "1")):
            concluidas.add(dom)

    abas = [n for n in wb.sheetnames if n not in ("Controle", "Instrucoes")]
    if len(abas) != len(dominios):
        raise SystemExit(f"ABORTADO: {len(abas)} abas contra {len(dominios)} linhas "
                         f"em `Controle`; a correspondencia por ordem nao e segura.")

    linhas = []
    for aba, sitio in zip(abas, dominios):
        corpo = aba.split(" ", 1)[1] if " " in aba else aba
        if not sitio.startswith(corpo.rstrip(".")):
            raise SystemExit(f"ABORTADO: a aba `{aba}` nao corresponde a `{sitio}`.")
        if concluidas and sitio not in concluidas:
            print(f"  ATENCAO: {sitio} nao consta como concluida; excluida do corpo")
            continue
        ws = wb[aba]
        cols = [c.value for c in ws[1]]
        idx = {v: cols.index(ROTULO[v]) for v in VARIAVEIS if ROTULO[v] in cols}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None:
                continue
            reg = {"site_id": sitio, "segmento_id": int(r[0]),
                   "n_caracteres": len(str(r[1] or "")), "texto": str(r[1] or "")}
            for v in VARIAVEIS:
                j = idx.get(v)
                reg[v] = 1 if (j is not None and r[j] not in (None, "")) else 0
            linhas.append(reg)

    out = REPO / args.out
    out.parent.mkdir(parents=True, exist_ok=True)
    campos = ["site_id", "segmento_id", "n_caracteres"] + VARIAVEIS + ["texto"]
    with out.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=campos, delimiter=";", quoting=csv.QUOTE_ALL)
        w.writeheader()
        w.writerows({k: l[k] for k in campos} for l in linhas)

    if not linhas:
        raise SystemExit("ABORTADO: nenhuma politica consta como concluida na aba "
                         "`Controle`. Marque `sim` a medida que cada uma for julgada.")
    sitios = sorted({l["site_id"] for l in linhas})
    print(f"corpo rotulado: {len(linhas):,} segmentos, {len(sitios)} politicas\n")
    print(f"  {'variavel':22}{'positivos':>11}{'documentos':>12}{'proporcao':>13}")
    for v in VARIAVEIS:
        p = sum(l[v] for l in linhas)
        d = len({l["site_id"] for l in linhas if l[v]})
        # Durante a rodada de marcacao ha politicas ainda nao concluidas, e uma
        # variavel pode nao ter positivo algum entre as concluidas.
        prop = f"1 : {int((len(linhas) - p) / p)}" if p else "sem positivos"
        print(f"  {ROTULO[v]:22}{p:>11}{d:>12}{prop:>13}")
    print(f"\n  {'politica':32}{'segmentos':>11}" + "".join(f"{ROTULO[v][:5]:>8}" for v in VARIAVEIS))
    c = Counter(l["site_id"] for l in linhas)
    for s in sorted(sitios, key=lambda x: -c[x]):
        pos = "".join(f"{sum(l[v] for l in linhas if l['site_id'] == s):>8}" for v in VARIAVEIS)
        print(f"  {s:32}{c[s]:>11}{pos}")
    print(f"\nsaida: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
