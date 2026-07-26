# -*- coding: utf-8 -*-
"""Monta o pacote de entrega ao segundo avaliador.

O calculo de concordancia so tem valor se o segundo avaliador trabalhar em
cegueira efetiva. Entregar a pasta de trabalho inteira comprometeria a medida por
dois caminhos: os rotulos do avaliador primario estao no arquivo principal, e os
pacotes de evidencia dos sitios fora do subconjunto permitiriam inferir criterios
por comparacao.

Esta rotina reune, portanto, exclusivamente: a planilha cega, o codebook e os
pacotes de evidencia dos 70 sitios sorteados. Nada mais.

Registre-se que os pacotes de evidencia reproduzem trechos de politicas de
privacidade e, com eles, contatos de encarregado. O segundo avaliador passa a
tratar dado pessoal e deve ser instruido a nao redistribuir o material nem
utiliza-lo para finalidade diversa da rotulagem.

Uso:
    python scripts/preparar_kappa.py --tcc "C:/caminho/para/TCC"
"""
from __future__ import annotations

import argparse
import shutil
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--tcc", required=True, help="pasta raiz do TCC")
    ap.add_argument("--planilha", default="Rotulagem/Kappa - Subconjunto Cego b9 (70 sitios).xlsx")
    ap.add_argument("--codebook", default="Codebook - Quatro Testes ML (canal + textuais) - v2.md")
    ap.add_argument("--destino", default="Rotulagem/entrega_segundo_avaliador")
    args = ap.parse_args()

    tcc = Path(args.tcc)
    destino = tcc / args.destino
    (destino / "evidencia").mkdir(parents=True, exist_ok=True)

    planilha = tcc / args.planilha
    wb = openpyxl.load_workbook(planilha)
    ws = wb["Rotulagem (cego)"]
    cols = [c.value for c in ws[1]]

    # conferencia de cegueira antes de qualquer copia
    for campo in ("status", "tem_canal_titular", "finalidade", "direitos_titular",
                  "transf_internacional"):
        if campo not in cols:
            raise SystemExit(f"ABORTADO: coluna {campo} ausente da planilha")
        j = cols.index(campo) + 1
        cheios = sum(1 for r in range(2, ws.max_row + 1)
                     if ws.cell(r, j).value not in (None, ""))
        if cheios:
            raise SystemExit(f"ABORTADO: {campo} tem {cheios} valores preenchidos; "
                             "a planilha nao esta cega")

    je = cols.index("evidencia_arquivo") + 1
    js = cols.index("site_id") + 1
    copiados = faltando = 0
    for r in range(2, ws.max_row + 1):
        rel = ws.cell(r, je).value
        origem = tcc / rel if rel else None
        if origem and origem.exists():
            shutil.copy2(origem, destino / "evidencia" / origem.name)
            copiados += 1
        else:
            print(f"  FALTA evidencia: {ws.cell(r, js).value}")
            faltando += 1

    # a planilha entregue aponta para a pasta local de evidencia
    for r in range(2, ws.max_row + 1):
        rel = ws.cell(r, je).value
        if rel:
            ws.cell(r, je).value = f"evidencia/{Path(rel).name}"
    wb.save(destino / planilha.name)
    shutil.copy2(tcc / args.codebook, destino / Path(args.codebook).name)

    print(f"\nplanilha cega conferida: nenhum rotulo preenchido")
    print(f"pacotes de evidencia copiados: {copiados}   ausentes: {faltando}")
    print(f"codebook incluido: {Path(args.codebook).name}")
    print(f"\ndestino: {destino}")
    print("Entregar esta pasta ao segundo avaliador. Nao incluir o arquivo principal")
    print("de rotulagem nem os demais pacotes de evidencia.")
    return 1 if faltando else 0


if __name__ == "__main__":
    raise SystemExit(main())
