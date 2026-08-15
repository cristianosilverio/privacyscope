# -*- coding: utf-8 -*-
"""Audita a consistencia interna da marcacao exaustiva.

ESCOPO E LIMITE
---------------
Esta rotina NAO julga se cada segmento satisfaz o construto: a decisao e semantica e
depende do codebook aplicado por leitor humano. Tentativas de reproduzi-la por
expressao regular foram avaliadas e descartadas — produziram centenas de divergencias
que, examinadas, revelaram-se erro do detector, e nao da marcacao.

Audita-se o que e verificavel sem juizo: a CONSISTENCIA. Segmentos de texto identico,
apos normalizacao, devem receber a mesma marcacao. Divergencia entre eles configura
deriva de criterio ao longo de uma tarefa extensa, executada em varias sessoes, e nao
depende de interpretacao para ser reconhecida como defeito.

Relata-se ainda, em separado, a citacao legal marcada como positiva — dispositivo
citado sem enunciado proprio nao declara finalidade nem direito, e a verificacao e
objetiva o bastante para constar.

Uso:
    python scripts/auditar_marcacao_completude.py --tcc "C:/caminho/TCC"
"""
from __future__ import annotations

import argparse
import os
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

VARIAVEIS = ["finalidade", "direitos_titular", "transf_internacional"]
ROTULO = {"finalidade": "Finalidade", "direitos_titular": "Direitos",
          "transf_internacional": "Transf. intern."}
MIN_COMPARAVEL = 25
CITACAO = re.compile(r"^(?:lei|decreto|resolucao|rdc|rcd|provimento|artigo)"
                     r"[^a-z]{0,4}[\d./º°-]", re.I)


def normaliza(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9 ]", " ", re.sub(r"\s+", " ", s).lower()).strip()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--tcc", default=os.environ.get("PRIVACYSCOPE_TCC"),
                    help="pasta raiz do TCC; na ausencia do argumento adota-se a variavel de ambiente PRIVACYSCOPE_TCC")
    ap.add_argument("--planilha", default="Rotulagem/Completude - 15 politicas.xlsx")
    ap.add_argument("--saida", default="Rotulagem/Auditoria da marcacao.xlsx")
    args = ap.parse_args()

    if not args.tcc:
        print("ERRO: a pasta do TCC nao foi informada.")
        print("  Informe --tcc, ou defina a variavel de ambiente:")
        print("    PowerShell:  $env:PRIVACYSCOPE_TCC = \"C:\\caminho\\TCC\"")
        print("    bash:        export PRIVACYSCOPE_TCC=/caminho/TCC")
        return 2

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(Path(args.tcc) / args.planilha, data_only=True)
    reg = defaultdict(lambda: defaultdict(list))
    citacoes = []
    total = 0
    for aba in wb.sheetnames:
        if aba in ("Controle", "Instrucoes"):
            continue
        ws = wb[aba]
        cols = [c.value for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None:
                continue
            total += 1
            texto = str(r[1] or "")
            chave = normaliza(texto)
            for v in VARIAVEIS:
                marcado = 1 if r[cols.index(ROTULO[v])] not in (None, "") else 0
                if marcado and CITACAO.match(chave):
                    citacoes.append((aba, r[0], v, texto))
                if len(chave) >= MIN_COMPARAVEL:
                    reg[v][chave].append((aba, int(r[0]), marcado, texto))

    linhas = []
    for v in VARIAVEIS:
        for chave, oc in reg[v].items():
            if len({m for _, _, m, _ in oc}) > 1:
                for aba, sid, m, texto in sorted(oc):
                    linhas.append({"variavel": ROTULO[v], "aba": aba, "segmento_id": sid,
                                   "marcado": m, "texto": texto})

    print(f"segmentos examinados: {total:,}\n")
    print("=== marcacao divergente entre segmentos de texto identico ===")
    for v in VARIAVEIS:
        d = [k for k, oc in reg[v].items() if len({m for _, _, m, _ in oc}) > 1]
        n = sum(len(reg[v][k]) for k in d)
        print(f"  {ROTULO[v]:20}{len(d):>4} texto(s), {n:>4} ocorrencia(s)")
    print(f"\n=== citacao legal marcada como positiva: {len(citacoes)} ===")
    for aba, sid, v, t in citacoes:
        print(f"  {aba:32} id={sid:<5} {ROTULO[v]:15} {' '.join(t.split())[:76]}")

    if not linhas and not citacoes:
        print("\n  nenhuma inconsistencia detectada.")
        return 0

    ws_out = openpyxl.Workbook()
    ws = ws_out.active
    ws.title = "Divergencias"
    cab = ["variavel", "aba", "segmento_id", "marcado", "texto", "decisao final"]
    ws.append(cab)
    for c in range(1, len(cab) + 1):
        ws.cell(1, c).font = Font(bold=True)
        ws.cell(1, c).fill = PatternFill("solid", fgColor="DDDDDD")
    for l in linhas:
        ws.append([l["variavel"], l["aba"], l["segmento_id"], l["marcado"], l["texto"], ""])
    for larg, col in ((16, "A"), (30, "B"), (12, "C"), (10, "D"), (100, "E"), (14, "F")):
        ws.column_dimensions[col].width = larg
    for linha in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        linha[0].alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    if citacoes:
        wc = ws_out.create_sheet("Citacao legal")
        wc.append(["aba", "segmento_id", "variavel", "texto", "decisao final"])
        for c in range(1, 6):
            wc.cell(1, c).font = Font(bold=True)
        for aba, sid, v, t in citacoes:
            wc.append([aba, sid, ROTULO[v], t, ""])
        for larg, col in ((30, "A"), (12, "B"), (16, "C"), (100, "D"), (14, "E")):
            wc.column_dimensions[col].width = larg

    destino = Path(args.tcc) / args.saida
    ws_out.save(destino)
    print(f"\nsaida: {destino}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
