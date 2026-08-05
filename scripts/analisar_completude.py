# -*- coding: utf-8 -*-
"""Confronta a marcacao exaustiva com a derivacao automatica dos rotulos.

O QUE SE MEDE
-------------
A derivacao de rotulos de segmento apoia-se em uma suposicao: a de que a passagem
transcrita durante a rotulagem esgota, ou quase esgota, o conteudo relevante do
documento. Esta rotina afere essa suposicao confrontando duas fontes sobre as
mesmas politicas: a marcacao exaustiva, em que o anotador percorreu todos os
segmentos, e a derivacao automatica, em que os segmentos foram assinalados por
sobreposicao com a passagem originalmente transcrita.

Duas grandezas resultam:

  REVOCACAO da derivacao — dos segmentos que a marcacao exaustiva reconheceu como
  relevantes, que fracao a derivacao automatica alcancou. E a medida direta da
  suposicao: revocacao elevada indica que a transcricao original era praticamente
  completa e que os segmentos nao rotulados podem ser tratados como negativos;
  revocacao baixa indica que o descarte por precaucao deve ser mantido.

  PRECISAO da derivacao — dos segmentos que a derivacao assinalou, que fracao a
  marcacao exaustiva confirma. Afere o efeito contrario: a sobreposicao posicional
  ocasionalmente arrasta para dentro da passagem material contiguo e alheio ao
  requisito, e essa parcela e ruido no conjunto de treino.

A rotina opera sobre planilha PARCIALMENTE preenchida, o que permite leitura
preliminar antes da conclusao da tarefa. Politicas ainda nao percorridas sao
identificadas pela coluna de conclusao da aba de Controle e excluidas do calculo:
uma politica sem marcacao alguma e indistinguivel, sem esse registro, de uma
politica percorrida em que nada se julgou relevante.

Uso:
    python scripts/analisar_completude.py --tcc "C:/caminho/TCC"
"""
from __future__ import annotations

import argparse
import csv
from math import sqrt
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
VARIAVEIS = ["finalidade", "direitos_titular", "transf_internacional"]
ROTULO = {"finalidade": "Finalidade", "direitos_titular": "Direitos",
          "transf_internacional": "Transf. intern."}


def wilson(k, n, z=1.959963984540054):
    if n == 0:
        return 0.0, 1.0
    f = k / n
    d = 1 + z * z / n
    c = (f + z * z / (2 * n)) / d
    h = z * sqrt(f * (1 - f) / n + z * z / (4 * n * n)) / d
    return max(0.0, c - h), min(1.0, c + h)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--tcc", required=True)
    ap.add_argument("--planilha", default="Rotulagem/Completude - 15 politicas.xlsx")
    ap.add_argument("--segmentos", default="outputs/segmentos_textuais.csv")
    ap.add_argument("--out", default="outputs/completude_resultados.csv")
    args = ap.parse_args()

    import openpyxl

    wb = openpyxl.load_workbook(Path(args.tcc) / args.planilha, data_only=True)

    # politicas concluidas, conforme registro na aba de Controle
    concluidas = set()
    if "Controle" in wb.sheetnames:
        ws = wb["Controle"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r and r[0] and r[3] and str(r[3]).strip().lower() in ("sim", "s", "x", "1"):
                concluidas.add(str(r[0]).strip())

    # marcacao exaustiva, por politica e variavel
    marcado = {}
    com_marca = set()
    for aba in wb.sheetnames:
        if aba in ("Controle", "Instrucoes"):
            continue
        sitio = aba.split(" ", 1)[1] if " " in aba else aba
        ws = wb[aba]
        cols = [c.value for c in ws[1]]
        idx = {ROTULO[v]: cols.index(ROTULO[v]) for v in VARIAVEIS if ROTULO[v] in cols}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None:
                continue
            sid = int(r[0])
            for v in VARIAVEIS:
                j = idx.get(ROTULO[v])
                if j is not None and r[j] not in (None, ""):
                    marcado[(sitio, v, sid)] = 1
                    com_marca.add(sitio)

    avaliadas = concluidas or com_marca
    if not concluidas and com_marca:
        print("AVISO: a aba de Controle nao registra politicas concluidas; adotam-se as")
        print("       que apresentam ao menos uma marcacao, o que subestima as politicas")
        print("       percorridas em que nada se julgou relevante.\n")

    with (REPO / args.segmentos).open(encoding="utf-8", newline="") as fh:
        S = [l for l in csv.DictReader(fh, delimiter=";") if l["site_id"] in avaliadas]

    print(f"politicas avaliadas: {len(avaliadas)} de 15")
    for s in sorted(avaliadas):
        print(f"  {s}")
    print()

    linhas = []
    print("=" * 96)
    print("CONFRONTO ENTRE A MARCACAO EXAUSTIVA E A DERIVACAO AUTOMATICA")
    print("=" * 96)
    print(f"  {'variavel':20}{'exaustiva':>11}{'derivada':>10}{'ambas':>8}"
          f"{'revocacao':>22}{'precisao':>20}")
    for v in VARIAVEIS:
        sub = [l for l in S if l["variavel"] == v]
        exa = {(l["site_id"], int(l["segmento_id"]))
               for l in sub if marcado.get((l["site_id"], v, int(l["segmento_id"])))}
        der = {(l["site_id"], int(l["segmento_id"])) for l in sub if l["y"] == "1"}
        amb = exa & der
        if not exa and not der:
            print(f"  {ROTULO[v]:20}{'—':>11}{'—':>10}{'—':>8}"
                  f"{'sem ocorrencia':>22}{'':>20}")
            continue
        rev = len(amb) / len(exa) if exa else float("nan")
        pre = len(amb) / len(der) if der else float("nan")
        rlo, rhi = wilson(len(amb), len(exa)) if exa else (0, 1)
        plo, phi = wilson(len(amb), len(der)) if der else (0, 1)
        print(f"  {ROTULO[v]:20}{len(exa):>11}{len(der):>10}{len(amb):>8}"
              f"{rev*100:>10.1f}% [{rlo*100:.0f}–{rhi*100:.0f}]"
              f"{pre*100:>10.1f}% [{plo*100:.0f}–{phi*100:.0f}]")
        linhas.append({"variavel": v, "exaustiva": len(exa), "derivada": len(der),
                       "ambas": len(amb), "revocacao": rev, "rev_inf": rlo, "rev_sup": rhi,
                       "precisao": pre, "prec_inf": plo, "prec_sup": phi})

    print("\n" + "=" * 96)
    print("LEITURA")
    print("=" * 96)
    for r in linhas:
        print(f"  {ROTULO[r['variavel']]}:")
        perdidos = r["exaustiva"] - r["ambas"]
        print(f"    a transcricao original nao alcancou {perdidos} segmento(s) que a marcacao")
        print(f"    exaustiva reconheceu como relevantes")
        if r["revocacao"] >= 0.85:
            print("    Revocacao elevada: a suposicao de completude se sustenta, e os segmentos")
            print("    nao rotulados podem ser convertidos em negativos, declarando-se a taxa.")
        elif r["revocacao"] >= 0.6:
            print("    Revocacao intermediaria: a conversao e defensavel se a taxa de falso")
            print("    negativo for declarada e incorporada a discussao de limitacoes.")
        else:
            print("    Revocacao baixa: convem manter o descarte por precaucao dos segmentos")
            print("    nao rotulados, sob pena de introduzir falso negativo sistematico.")
        excedentes = r["derivada"] - r["ambas"]
        if excedentes:
            print(f"    a derivacao assinalou {excedentes} segmento(s) que a marcacao exaustiva")
            print("    nao confirma — material contiguo arrastado pela sobreposicao posicional")

    if len(avaliadas) < 15:
        print(f"\n  LEITURA PRELIMINAR: {len(avaliadas)} de 15 politicas. Os intervalos")
        print("  sao amplos e as estimativas pontuais nao devem orientar decisao definitiva.")

    if linhas:
        out = REPO / args.out
        out.parent.mkdir(parents=True, exist_ok=True)
        with out.open("w", encoding="utf-8", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=list(linhas[0].keys()), delimiter=";")
            w.writeheader(); w.writerows(linhas)
        print(f"\nsaida: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
