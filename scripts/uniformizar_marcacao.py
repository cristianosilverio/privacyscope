# -*- coding: utf-8 -*-
"""Uniformiza a marcacao de segmentos de texto identico.

A auditoria de consistencia relaciona os segmentos que, tendo texto identico apos
normalizacao, receberam marcacao divergente. A divergencia decorre de deriva de
criterio ao longo de tarefa extensa, executada em varias sessoes, e nao de diferenca
entre os segmentos — que sao, por construcao, o mesmo texto.

Uniformiza-se pelo valor decidido pelo anotador apos revisao da auditoria. O sentido
da uniformizacao e parametro explicito, e nao presuncao da rotina: adotar sempre o
valor positivo seria conveniente e arbitrario, e a decisao cabe a quem aplicou o
codebook.

A rotina relaciona o que sera alterado antes de gravar, e exige confirmacao.

Uso:
    python scripts/uniformizar_marcacao.py --tcc "C:/caminho/TCC"
    python scripts/uniformizar_marcacao.py --tcc "C:/caminho/TCC" --para 1 --aplicar
"""
from __future__ import annotations

import argparse
import os
import re
import shutil
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

VARIAVEIS = ["finalidade", "direitos_titular", "transf_internacional"]
ROTULO = {"finalidade": "Finalidade", "direitos_titular": "Direitos",
          "transf_internacional": "Transf. intern."}
MIN_COMPARAVEL = 25


def normaliza(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9 ]", " ", re.sub(r"\s+", " ", s).lower()).strip()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--tcc", default=os.environ.get("PRIVACYSCOPE_TCC"),
                    help="pasta raiz do TCC; na ausencia do argumento adota-se a variavel de ambiente PRIVACYSCOPE_TCC")
    ap.add_argument("--planilha", default="Rotulagem/Completude - 15 politicas.xlsx")
    ap.add_argument("--para", type=int, choices=(0, 1), default=1,
                    help="valor para o qual uniformizar os casos divergentes")
    ap.add_argument("--aplicar", action="store_true")
    args = ap.parse_args()

    if not args.tcc:
        print("ERRO: a pasta do TCC nao foi informada.")
        print("  Informe --tcc, ou defina a variavel de ambiente:")
        print("    PowerShell:  $env:PRIVACYSCOPE_TCC = \"C:\\caminho\\TCC\"")
        print("    bash:        export PRIVACYSCOPE_TCC=/caminho/TCC")
        return 2

    import openpyxl

    caminho = Path(args.tcc) / args.planilha
    wb = openpyxl.load_workbook(caminho)

    # levanta os textos divergentes por variavel
    reg = defaultdict(lambda: defaultdict(list))
    for aba in wb.sheetnames:
        if aba in ("Controle", "Instrucoes"):
            continue
        ws = wb[aba]
        cols = [c.value for c in ws[1]]
        for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
            if r[0].value is None:
                continue
            chave = normaliza(r[1].value)
            if len(chave) < MIN_COMPARAVEL:
                continue
            for v in VARIAVEIS:
                j = cols.index(ROTULO[v])
                marcado = 1 if r[j].value not in (None, "") else 0
                reg[v][chave].append((aba, i, j + 1, marcado, r[1].value))

    alterar = []
    for v in VARIAVEIS:
        for chave, oc in reg[v].items():
            if len({m for _, _, _, m, _ in oc}) > 1:
                for aba, linha, col, m, texto in oc:
                    if m != args.para:
                        alterar.append((v, aba, linha, col, m, texto))

    print(f"uniformizacao para o valor: {args.para}\n")
    print(f"celulas a alterar: {len(alterar)}\n")
    por_var = defaultdict(int)
    for v, *_ in alterar:
        por_var[v] += 1
    for v in VARIAVEIS:
        print(f"  {ROTULO[v]:20}{por_var[v]:>4} celula(s)")
    print()
    for v, aba, linha, col, m, texto in alterar[:12]:
        print(f"  {aba:30} linha {linha:<5} {ROTULO[v]:15} {m} -> {args.para}")
        print(f"     {' '.join(str(texto).split())[:96]}")
    if len(alterar) > 12:
        print(f"  ... e mais {len(alterar) - 12}")

    if not args.aplicar:
        print("\n  modo de relacao: nada foi gravado. Use --aplicar para efetivar.")
        return 0
    if not alterar:
        print("\n  nada a fazer.")
        return 0

    backup = caminho.with_suffix(f".bak_{datetime.now():%Y%m%d%H%M}.xlsx")
    shutil.copy2(caminho, backup)
    for v, aba, linha, col, m, texto in alterar:
        wb[aba].cell(linha, col).value = args.para if args.para == 1 else None
    wb.save(caminho)
    print(f"\n  {len(alterar)} celula(s) alterada(s); copia de seguranca em {backup.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
